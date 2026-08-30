#!/usr/bin/env python3
"""
健康数据合并脚本
读取快捷指令导出的 JSON → 验证 → 合并到 health_data.xlsx

用法:
  python3 merge_health_data.py <JSON文件路径> [--excel <Excel路径>] [--dry-run]

示例:
  # 只验证不写入（推荐先跑一次看看数据对不对）
  python3 merge_health_data.py ~/Downloads/health_export.json --dry-run

  # 验证 + 写入 Excel
  python3 merge_health_data.py ~/Downloads/health_export.json

  # 指定 Excel 路径
  python3 merge_health_data.py ~/Downloads/health_export.json --excel ./data/health_data.xlsx
"""

import json
import sys
import os
import re
import shutil
import argparse
from datetime import datetime, timedelta, timezone
from collections import defaultdict
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("需要安装 openpyxl: pip install openpyxl")
    sys.exit(1)


# ============================================================
# 配置 (与 health-tracker-v2/src/lib/excel.ts 保持一致)
# ============================================================

SLEEP_SHEET = "睡眠记录"
SLEEP_HEADERS = ["日期", "睡眠时长(小时)", "卧床时长(小时)", "数据来源", "备注", "同步时间"]

ACTIVITY_SHEET = "活动卡路里"
ACTIVITY_HEADERS = ["日期", "活动卡路里(kcal)", "数据来源", "同步时间"]

STRENGTH_SHEET = "力量训练"
STRENGTH_HEADERS = ["日期", "训练类型", "卡路里(kcal)", "疲劳度(RPE)", "备注"]

CARDIO_SHEET = "有氧训练"
CARDIO_HEADERS = ["日期", "训练类型", "卡路里(kcal)", "疲劳度(RPE)", "备注"]

WEIGHT_SHEET = "体重记录"
WEIGHT_HEADERS = ["日期", "体重(kg)", "数据来源", "备注", "同步时间"]

BODYFAT_SHEET = "体脂率记录"
BODYFAT_HEADERS = ["日期", "体脂率(%)", "数据来源", "备注", "同步时间"]

BJ_TZ = timezone(timedelta(hours=8))

# 合理范围 (用于剔除脏数据)
WEIGHT_RANGE = (20, 200)      # kg
BODYFAT_RANGE = (3, 60)       # %

# 睡眠分期: 计入睡眠时长的 stage key
ASLEEP_STAGES = {"core", "deep", "rem", "unspecified"}
INBED_STAGES = {"inbed"}
AWAKE_STAGES = {"awake"}

# 中文分期 → 标准 stage key
CN_STAGE_MAP = {
    "核心睡眠": "core", "浅睡": "core", "核心": "core",
    "深度睡眠": "deep", "深睡": "deep", "深度": "deep",
    "快速眼动": "rem",
    "入睡": "unspecified", "未分级": "unspecified", "未分类": "unspecified",
    "在床上": "inbed", "在床": "inbed", "卧床": "inbed",
    "清醒": "awake", "醒来": "awake",
}


def norm_sleep_stage(value):
    """
    归一化睡眠分期值 → stage key (core/deep/rem/unspecified/inbed/awake)

    兼容三种形态:
      - 完整标识符: HKCategoryValueSleepAnalysisAsleepCore / AsleepCore
      - 英文简写:   Core / Deep / REM / Asleep / InBed / Awake
      - 中文:       核心睡眠 / 深度睡眠 / REM / 入睡 / 在床上 / 清醒
    """
    s = str(value).strip()
    if not s:
        return ""

    # 中文直接查表
    if s in CN_STAGE_MAP:
        return CN_STAGE_MAP[s]
    for cn, key in CN_STAGE_MAP.items():
        if cn in s:
            return key

    t = s.lower().replace("_", "").replace("-", "").replace(" ", "")

    # 去掉完整标识符前缀
    for prefix in ("hkcategoryvaluesleepanalysis", "asleep"):
        if t.startswith(prefix):
            t = t[len(prefix):]

    # "Asleep" 去掉前缀后为空 → 未分级睡眠 (计入睡眠)
    if not t:
        return "unspecified"

    if t in ("core", "deep", "rem", "unspecified", "inbed", "awake"):
        return t
    if t == "asleepunspecified":
        return "unspecified"

    return t  # 未知值原样返回, 由调用方判断

# 锻炼类型分类
STRENGTH_KEYWORDS = ["力量", "strength", "weight", "举重", "杠铃", "哑铃"]
CARDIO_KEYWORDS = [
    "跑", "run", "骑行", "cycl", "步行", "walk", "游泳", "swim",
    "瑜伽", "yoga", "有氧", "cardio", "hiit", "椭圆", "划船", "dance",
]


# ============================================================
# 日期解析
# ============================================================

def parse_date(s):
    """解析多种日期格式, 返回 datetime(带时区)"""
    if isinstance(s, datetime):
        d = s
        return d if d.tzinfo else d.replace(tzinfo=BJ_TZ)

    s = str(s).strip()
    if not s:
        return None

    # 中文日期格式: 2026年8月28日 上午12:51 / 2026年8月28日 下午11:25
    if "年" in s and "月" in s:
        cn = _parse_chinese_date(s)
        if cn:
            return cn

    # 统一 +0800 → +08:00
    s = s.replace("+0800", "+08:00").replace("+0900", "+09:00")

    formats = [
        "%Y-%m-%d %H:%M:%S%z",       # 2026-08-03 23:15:00 +08:00
        "%Y-%m-%dT%H:%M:%S%z",       # 2026-08-03T23:15:00+08:00
        "%Y-%m-%d %H:%M:%S",         # 2026-08-03 23:15:00
        "%Y-%m-%dT%H:%M:%S",         # 2026-08-03T23:15:00
        "%Y-%m-%d %H:%M",            # 2026-08-03 23:15
        "%Y-%m-%d",                  # 2026-08-03
    ]
    for fmt in formats:
        try:
            dt = datetime.strptime(s, fmt)
            return dt if dt.tzinfo else dt.replace(tzinfo=BJ_TZ)
        except ValueError:
            continue

    try:
        dt = datetime.fromisoformat(s)
        return dt if dt.tzinfo else dt.replace(tzinfo=BJ_TZ)
    except Exception:
        pass

    return None


def _parse_chinese_date(s):
    """解析中文日期: 2026年8月28日 上午12:51 / 下午11:25"""
    m = re.match(
        r"(\d{4})年(\d{1,2})月(\d{1,2})日"
        r"(?:\s*(上午|下午|凌晨|中午|晚上)?\s*(\d{1,2}):(\d{2})(?::(\d{2}))?)?",
        s,
    )
    if not m:
        return None

    year, month, day = int(m.group(1)), int(m.group(2)), int(m.group(3))
    ampm, hour_s, minute_s, sec_s = m.group(4), m.group(5), m.group(6), m.group(7)

    hour = int(hour_s) if hour_s else 0
    minute = int(minute_s) if minute_s else 0
    second = int(sec_s) if sec_s else 0

    # 中文 12 小时制 → 24 小时制
    # 注意: 中文习惯里"上午12:51"实际是凌晨 00:51
    if ampm in ("下午", "晚上") and hour < 12:
        hour += 12
    elif ampm == "上午" and hour == 12:
        hour = 0
    elif ampm == "中午" and hour < 12:
        hour += 12

    try:
        return datetime(year, month, day, hour, minute, second, tzinfo=BJ_TZ)
    except ValueError:
        return None


def bj_date_str(dt):
    """datetime → 北京时间 'YYYY-MM-DD'"""
    if not dt:
        return ""
    return dt.astimezone(BJ_TZ).strftime("%Y-%m-%d")


def normalize_samples(raw):
    """
    兼容快捷指令的多种输出形态:
      - list            → 直接返回 (标准数组)
      - str (JSON 数组)  → json.loads
      - str (JSONL)     → 按行 json.loads, 每行一个 JSON 对象
      - None / 空        → []
    """
    if raw is None:
        return []
    if isinstance(raw, list):
        return [x for x in raw if isinstance(x, dict)]
    if not isinstance(raw, str):
        return []

    s = raw.strip()
    if not s:
        return []

    # 先试完整 JSON 数组
    try:
        parsed = json.loads(s)
        if isinstance(parsed, list):
            return [x for x in parsed if isinstance(x, dict)]
        if isinstance(parsed, dict):
            return [parsed]
    except json.JSONDecodeError:
        pass

    # 再试 JSONL: 每行一个 JSON 对象
    out = []
    for line in s.split("\n"):
        line = line.strip().rstrip(",")
        if not line:
            continue
        try:
            obj = json.loads(line)
            if isinstance(obj, dict):
                out.append(obj)
            elif isinstance(obj, list):
                out.extend([x for x in obj if isinstance(x, dict)])
        except json.JSONDecodeError:
            continue
    return out


def find_samples(data, *names):
    """
    按多个可能的字段名查找样本数组, 容错命名差异。
    先精确匹配, 再宽松匹配 (忽略大小写/下划线)。
    返回 (样本列表, 命中的字段名 或 None)
    """
    for n in names:
        s = normalize_samples(data.get(n))
        if s:
            return s, n

    norm = lambda x: x.lower().replace("_", "").replace("-", "")
    targets = [norm(n) for n in names]
    for k, v in data.items():
        kl = norm(k)
        for t in targets:
            if t and (t in kl or kl in t):
                s = normalize_samples(v)
                if s:
                    return s, k
    return [], None


def sleep_date(dt):
    """睡眠日期归属: 中午12点前归前一晚"""
    bj = dt.astimezone(BJ_TZ)
    if bj.hour < 12:
        return (bj - timedelta(days=1)).strftime("%Y-%m-%d")
    return bj.strftime("%Y-%m-%d")


SOURCE_PLACEHOLDERS = {"健康样本", "health sample", "重复项目", "repeat item", "项目"}


def clean_source(source):
    """清理来源: 把快捷指令默认占位符当成空"""
    s = str(source).strip()
    low = s.lower()
    if low in {p.lower() for p in SOURCE_PLACEHOLDERS}:
        return ""
    return s


def is_watch(source):
    """判断来源是否包含 Watch"""
    s = clean_source(source).lower()
    return "watch" in s or "apple" in s


# ============================================================
# 睡眠解析
# ============================================================

def _fill_missing_end(samples, max_gap_min=240):
    """
    容错: 若某条缺有效 end (快捷指令里 end 选错成"值"时常见),
    用下一条样本的 start 推算。睡眠分期是连续的, 相邻起点即上一段终点。
    原地修改, 不返回。

    max_gap_min: 推算间隔上限(分钟)。超过则视为跨段(夜晚之间的白天间隔),
    不做推算——宁可丢掉最后一小段, 也不能把白天算成睡眠。
    """
    idx = sorted(
        range(len(samples)),
        key=lambda i: parse_date(samples[i].get("start", "")) or datetime.max.replace(tzinfo=BJ_TZ),
    )
    for n, i in enumerate(idx):
        s = samples[i]
        start = parse_date(s.get("start", s.get("startDate", "")))
        end = parse_date(s.get("end", s.get("endDate", "")))
        if not start:
            continue
        # end 有效且不早于 start → 无需修补
        if end and end > start:
            continue
        # 取下一条的 start
        if n + 1 < len(idx):
            nxt = parse_date(samples[idx[n + 1]].get("start", ""))
            if not nxt or nxt <= start:
                continue
            gap = (nxt - start).total_seconds() / 60
            if gap > max_gap_min:
                continue  # 跨段间隔, 放弃推算
            s["_end_inferred"] = True
            s["end"] = nxt.strftime("%Y-%m-%d %H:%M:%S")


def parse_sleep(samples):
    """解析睡眠样本, 按天聚合"""
    daily = defaultdict(lambda: {
        "sleep_min": 0, "bed_min": 0, "source": "", "count": 0,
        "deep": 0, "rem": 0, "core": 0, "unspecified": 0,
    })
    skipped = 0

    # end 缺失/错值 → 用下一条 start 推算
    _fill_missing_end(samples)
    inferred = sum(1 for s in samples if s.get("_end_inferred"))

    # 容错: 若所有样本 source 均为空/占位符, 放宽来源过滤
    any_source = any(
        clean_source(s.get("source", s.get("sourceName", ""))) for s in samples
    )

    for s in samples:
        start = parse_date(s.get("start", s.get("startDate", "")))
        end = parse_date(s.get("end", s.get("endDate", "")))
        value = str(s.get("value", "")).strip()
        source = clean_source(s.get("source", s.get("sourceName", "")))

        if not start or not end:
            skipped += 1
            continue
        if any_source and not is_watch(source):
            skipped += 1
            continue

        minutes = (end - start).total_seconds() / 60
        if minutes <= 0 or minutes > 24 * 60:
            skipped += 1
            continue

        date = sleep_date(start)
        stage = norm_sleep_stage(value)

        if stage in ASLEEP_STAGES:
            daily[date]["sleep_min"] += minutes
            daily[date]["count"] += 1
            daily[date][stage] += minutes
        elif stage in INBED_STAGES:
            daily[date]["bed_min"] += minutes
        else:
            # 未知分期 (含 awake) → 不计入
            pass

        if not daily[date]["source"]:
            daily[date]["source"] = source

    result = {}
    for date, d in sorted(daily.items()):
        result[date] = {
            "sleep_hours": round(d["sleep_min"] / 60, 1),
            "bed_hours": round(d["bed_min"] / 60, 1),
            "source": "Apple Health",
            "stages": {
                "deep": round(d["deep"]),
                "rem": round(d["rem"]),
                "core": round(d["core"]),
                "unspec": round(d["unspecified"]),
            },
            "samples": d["count"],
        }
    return result, skipped


# ============================================================
# 活动能量解析
# ============================================================

def parse_activity(samples):
    """解析活动能量样本, 按天求和"""
    daily = defaultdict(lambda: {"cal": 0.0, "count": 0})
    skipped = 0

    any_source = any(
        clean_source(s.get("source", s.get("sourceName", ""))) for s in samples
    )

    # 自适应异常值上限:
    #   按天/按小时分组 → 每条是当日(或每小时)汇总, 上限放宽到 5000
    #   原始样本(每分钟一条) → 单条不可能太高, 上限 1000 用于剔除脏数据
    max_val = 5000 if len(samples) <= 60 else 1000

    for s in samples:
        start = parse_date(s.get("start", s.get("startDate", "")))
        source = clean_source(s.get("source", s.get("sourceName", "")))

        if not start:
            skipped += 1
            continue
        if any_source and not is_watch(source):
            skipped += 1
            continue

        try:
            val = float(s.get("value", 0))
        except (ValueError, TypeError):
            skipped += 1
            continue

        # val <= 0 视为"当天无数据"(常见于导出当天尚未过完), 跳过而非写入 0
        if val <= 0 or val > max_val:
            skipped += 1
            continue

        date = bj_date_str(start)
        daily[date]["cal"] += val
        daily[date]["count"] += 1

    result = {}
    for date, d in sorted(daily.items()):
        result[date] = {
            "calories": round(d["cal"]),
            "source": "Apple Health",
            "samples": d["count"],
        }
    return result, skipped


# ============================================================
# 体重 / 体脂率解析
# ============================================================

def parse_measure(samples, val_range):
    """
    解析体重或体脂率: 同一天多条记录取时间最晚的一条。
    返回 ({date: {"value": float, "source": str}}, skipped)
    """
    latest = {}  # date -> (datetime, value, source)
    skipped = 0

    # 体重/体脂来源多样 (体脂秤 / 手动录入 / 第三方 App),
    # 不像睡眠那样只认 Watch, 这里不做来源过滤

    lo, hi = val_range
    for s in samples:
        start = parse_date(s.get("start", s.get("startDate", "")))
        source = clean_source(s.get("source", s.get("sourceName", "")))

        if not start:
            skipped += 1
            continue

        try:
            val = float(s.get("value", 0))
        except (ValueError, TypeError):
            skipped += 1
            continue

        if val < lo or val > hi:
            skipped += 1
            continue

        date = bj_date_str(start)
        if date not in latest or start > latest[date][0]:
            latest[date] = (start, val, source or "Apple Health")

    result = {
        d: {"value": round(v, 1), "source": src}
        for d, (_, v, src) in sorted(latest.items())
    }
    return result, skipped


# ============================================================
# 锻炼记录解析
# ============================================================

def classify(type_str):
    """判断锻炼类型: strength / cardio"""
    t = str(type_str).lower()
    for kw in STRENGTH_KEYWORDS:
        if kw in t:
            return "strength"
    for kw in CARDIO_KEYWORDS:
        if kw in t:
            return "cardio"
    return "cardio"  # 默认归有氧


def parse_workouts(workouts):
    """解析锻炼记录"""
    strength, cardio = [], []
    skipped = 0

    any_source = any(
        clean_source(w.get("source", w.get("sourceName", ""))) for w in workouts
    )

    for w in workouts:
        start = parse_date(w.get("start", w.get("startDate", "")))
        source = clean_source(w.get("source", w.get("sourceName", "")))
        wtype = str(w.get("type", w.get("activityType", ""))).strip()

        if not start:
            skipped += 1
            continue
        if any_source and not is_watch(source):
            skipped += 1
            continue

        try:
            energy = float(w.get("energy", w.get("totalEnergyBurned", 0)))
        except (ValueError, TypeError):
            energy = 0

        date = bj_date_str(start)
        record = {
            "date": date,
            "type": wtype,
            "calories": round(energy),
            "source": source,
        }
        cat = classify(wtype)
        if cat == "strength":
            strength.append(record)
        else:
            cardio.append(record)

    return {"strength": strength, "cardio": cardio}, skipped


# ============================================================
# 数据验证
# ============================================================

def validate(sleep, activity, workouts, weight=None, bodyfat=None):
    warnings = []
    weight = weight or {}
    bodyfat = bodyfat or {}

    # 睡眠
    for date, d in sleep.items():
        h = d["sleep_hours"]
        if h == 0:
            warnings.append(f"{date} 睡眠时长为 0")
        elif h > 14:
            warnings.append(f"{date} 睡眠 {h}h 异常偏大")
        elif 0 < h < 2:
            warnings.append(f"{date} 睡眠 {h}h 偏小 (可能只有午休)")

    # 活动能量
    for date, d in activity.items():
        c = d["calories"]
        if c > 2000:
            warnings.append(f"{date} 活动能量 {c}kcal 异常偏大")
        elif c < 50:
            warnings.append(f"{date} 活动能量 {c}kcal 偏低")

    # 锻炼
    for r in workouts["strength"] + workouts["cardio"]:
        if r["calories"] > 1000:
            warnings.append(f"{r['date']} 锻炼能量 {r['calories']}kcal 异常偏大")

    # 缺失日期 (体重/体脂非每日必测, 不单独报缺失)
    all_dates = sorted(
        set(sleep.keys()) | set(activity.keys())
        | set(weight.keys()) | set(bodyfat.keys())
    )
    if all_dates:
        cur = datetime.strptime(all_dates[0], "%Y-%m-%d")
        end = datetime.strptime(all_dates[-1], "%Y-%m-%d")
        while cur <= end:
            ds = cur.strftime("%Y-%m-%d")
            if ds not in sleep:
                warnings.append(f"{ds} 无睡眠数据")
            if ds not in activity:
                warnings.append(f"{ds} 无活动能量数据")
            cur += timedelta(days=1)

    return warnings


# ============================================================
# Excel 操作
# ============================================================

def find_excel(custom=None):
    if custom and os.path.exists(custom):
        return custom
    candidates = [
        Path.cwd() / "data" / "health_data.xlsx",
        Path(__file__).resolve().parent.parent / "data" / "health_data.xlsx",
        Path.home() / "Library" / "CloudStorage" / "OneDrive-个人" / "projects" / "health_agent" / "health-tracker-v2" / "data" / "health_data.xlsx",
    ]
    for p in candidates:
        if p.exists():
            return str(p)
    print("找不到 health_data.xlsx, 请用 --excel 指定路径")
    for p in candidates:
        print(f"  尝试过: {p}")
    sys.exit(1)


def _val_differs(old, new):
    """判断两个值是否有实质差异 (数值容差 0.05)"""
    if old is None and new in (None, ""):
        return False
    if old is None:
        return True
    if new is None:
        return False
    try:
        return abs(float(old) - float(new)) > 0.05
    except (ValueError, TypeError):
        pass
    return str(old).strip() != str(new).strip()


def upsert_by_date(wb, sheet_name, headers, rows):
    """
    按日期 upsert (已存在则覆盖, 否则追加)
    返回 (写入条数, 新增日期列表, 变更明细列表)
    变更明细: [{"date","field","old","new"}, ...]
    """
    if sheet_name not in wb.sheetnames:
        print(f"  Sheet '{sheet_name}' 不存在, 跳过")
        return 0, [], []
    ws = wb[sheet_name]

    # 读取现有数据
    existing = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        existing.append(list(row))
    existing_dates = {str(r[0]) for r in existing}

    added, changed = [], []

    for new_row in rows:
        date_val = str(new_row.get("日期", ""))
        if date_val in existing_dates:
            # 更新
            for i, r in enumerate(existing):
                if str(r[0]) == date_val:
                    for j, h in enumerate(headers):
                        if h not in new_row:
                            continue
                        # 同步时间每次都刷新, 不计入变更
                        if h == "同步时间":
                            r[j] = new_row[h]
                            continue
                        if _val_differs(r[j], new_row[h]):
                            changed.append({
                                "date": date_val,
                                "field": h,
                                "old": r[j],
                                "new": new_row[h],
                            })
                        r[j] = new_row[h]
                    # 写回 Excel
                    for j in range(len(headers)):
                        ws.cell(row=i + 2, column=j + 1, value=r[j])
                    break
        else:
            # 追加
            ws.append([new_row.get(h, "") for h in headers])
            existing.append([new_row.get(h, "") for h in headers])
            existing_dates.add(date_val)
            added.append(date_val)

    return len(rows), added, changed


def append_dedup(wb, sheet_name, headers, rows, dedup_keys=("日期", "训练类型")):
    """追加锻炼记录, 按 dedup_keys 去重, 返回 (新增条数, 新增描述列表)"""
    if sheet_name not in wb.sheetnames:
        print(f"  Sheet '{sheet_name}' 不存在, 跳过")
        return 0, []
    ws = wb[sheet_name]

    # 读取已有记录的 dedup 指纹
    existing_keys = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        key = tuple(str(row[headers.index(k)]) if headers.index(k) < len(row) else "" for k in dedup_keys)
        existing_keys.add(key)

    count = 0
    added = []
    for new_row in rows:
        key = tuple(str(new_row.get(k, "")) for k in dedup_keys)
        if key in existing_keys:
            continue
        ws.append([new_row.get(h, "") for h in headers])
        existing_keys.add(key)
        added.append(f"{new_row.get('日期', '')} {new_row.get('训练类型', '')}".strip())
        count += 1
    return count, added


def print_change_report(report):
    """输出变更报告: 区分新增 与 Apple 数据修正"""
    total_added = sum(len(a) for _, a, _ in report)
    total_changed = sum(len(c) for _, _, c in report)

    if not total_added and not total_changed:
        print(f"\n变更报告: 无变化 (与 Excel 已有数据完全一致)")
        return

    print(f"\n变更报告:")
    for sheet, added, changed in report:
        if not added and not changed:
            continue
        print(f"  [{sheet}]")
        for d in added:
            print(f"    + {d}  新增")
        for c in changed:
            print(f"    ~ {c['date']}  {c['field']}: {c['old']} -> {c['new']}")

    print(f"\n  合计: 新增 {total_added} 条, 修正 {total_changed} 处")


def merge_to_excel(excel_path, sleep, activity, workouts,
                    weight=None, bodyfat=None):
    weight = weight or {}
    bodyfat = bodyfat or {}
    print(f"\n写入 Excel: {excel_path}")
    backup = excel_path + ".bak"
    shutil.copy2(excel_path, backup)
    print(f"已备份: {backup}")

    wb = load_workbook(excel_path)
    now = datetime.now(BJ_TZ).isoformat()

    report = []  # [(sheet名, 新增日期列表, 变更明细列表)]

    # 睡眠
    sleep_rows = []
    for date, d in sorted(sleep.items()):
        sleep_rows.append({
            "日期": date,
            "睡眠时长(小时)": d["sleep_hours"],
            "卧床时长(小时)": d["bed_hours"],
            "数据来源": "Apple Health",
            "备注": "",
            "同步时间": now,
        })
    if sleep_rows:
        n, added, changed = upsert_by_date(wb, SLEEP_SHEET, SLEEP_HEADERS, sleep_rows)
        print(f"  睡眠记录: {n} 条")
        report.append((SLEEP_SHEET, added, changed))

    # 活动能量
    act_rows = []
    for date, d in sorted(activity.items()):
        act_rows.append({
            "日期": date,
            "活动卡路里(kcal)": d["calories"],
            "数据来源": "Apple Health",
            "同步时间": now,
        })
    if act_rows:
        n, added, changed = upsert_by_date(wb, ACTIVITY_SHEET, ACTIVITY_HEADERS, act_rows)
        print(f"  活动卡路里: {n} 条")
        report.append((ACTIVITY_SHEET, added, changed))

    # 体重
    wt_rows = []
    for date, d in sorted(weight.items()):
        wt_rows.append({
            "日期": date,
            "体重(kg)": d["value"],
            "数据来源": d["source"] or "Apple Health",
            "备注": "",
            "同步时间": now,
        })
    if wt_rows:
        n, added, changed = upsert_by_date(wb, WEIGHT_SHEET, WEIGHT_HEADERS, wt_rows)
        print(f"  体重记录: {n} 条")
        report.append((WEIGHT_SHEET, added, changed))

    # 体脂率
    bf_rows = []
    for date, d in sorted(bodyfat.items()):
        bf_rows.append({
            "日期": date,
            "体脂率(%)": d["value"],
            "数据来源": d["source"] or "Apple Health",
            "备注": "",
            "同步时间": now,
        })
    if bf_rows:
        n, added, changed = upsert_by_date(wb, BODYFAT_SHEET, BODYFAT_HEADERS, bf_rows)
        print(f"  体脂率记录: {n} 条")
        report.append((BODYFAT_SHEET, added, changed))

    # 力量训练
    str_rows = []
    for r in workouts["strength"]:
        str_rows.append({
            "日期": r["date"],
            "训练类型": r["type"],
            "卡路里(kcal)": r["calories"],
            "疲劳度(RPE)": "",
            "备注": "",
        })
    if str_rows:
        n, added = append_dedup(wb, STRENGTH_SHEET, STRENGTH_HEADERS, str_rows)
        print(f"  力量训练: {n} 条 (新增)")
        report.append((STRENGTH_SHEET, added, []))

    # 有氧训练
    car_rows = []
    for r in workouts["cardio"]:
        car_rows.append({
            "日期": r["date"],
            "训练类型": r["type"],
            "卡路里(kcal)": r["calories"],
            "疲劳度(RPE)": "",
            "备注": "",
        })
    if car_rows:
        n, added = append_dedup(wb, CARDIO_SHEET, CARDIO_HEADERS, car_rows)
        print(f"  有氧训练: {n} 条 (新增)")
        report.append((CARDIO_SHEET, added, []))

    wb.save(excel_path)
    print(f"Excel 已更新")

    print_change_report(report)
    return report


# ============================================================
# 主函数
# ============================================================

def main():
    parser = argparse.ArgumentParser(description="健康数据合并脚本")
    parser.add_argument("json_file", help="快捷指令导出的 JSON 文件路径")
    parser.add_argument("--excel", help="Excel 文件路径 (默认自动查找)")
    parser.add_argument("--dry-run", action="store_true", help="只验证不写入 Excel")
    parser.add_argument(
        "--include-today", action="store_true",
        help="保留导出当天的数据 (默认剔除, 因为当天未过完数据不完整)",
    )
    args = parser.parse_args()

    if not os.path.exists(args.json_file):
        print(f"JSON 文件不存在: {args.json_file}")
        sys.exit(1)

    print(f"读取: {args.json_file}")
    with open(args.json_file, "r", encoding="utf-8") as f:
        data = json.load(f)

    print(f"导出时间: {data.get('exportDate', '?')}")
    print(f"天数: {data.get('days', '?')}")

    # 字段名容错: activty/bodyfat/weightSamples 等各种命名都能识别
    sleep_samples = normalize_samples(data.get("sleep"))
    activity_samples = normalize_samples(
        data.get("activity") or data.get("activty")
    )
    workouts = normalize_samples(data.get("workouts"))
    weight_samples, w_key = find_samples(
        data, "weight", "bodyMass", "weightSamples"
    )
    bodyfat_samples, bf_key = find_samples(
        data, "bodyFat", "bodyFatPercentage", "bodyfat", "fatPercentage"
    )

    if not activity_samples:
        for k, v in data.items():
            if k.lower().startswith("activ"):
                activity_samples = normalize_samples(v)
                if activity_samples:
                    print(f"  提示: 使用了字段 '{k}' 作为活动能量数据")
                    break

    if "activity" not in data and "activty" not in data:
        print("  [注意] JSON 中没有 activity/activty 字段, 快捷指令未导出活动能量数据")
    elif not activity_samples:
        print("  [注意] 活动能量字段存在, 但为空数组")

    if not weight_samples:
        print("  [注意] 未找到体重数据 (字段可为 weight / bodyMass)")
    if not bodyfat_samples:
        print("  [注意] 未找到体脂率数据 (字段可为 bodyFat / bodyFatPercentage)")

    print(f"\n原始样本:")
    print(f"  睡眠: {len(sleep_samples)} 条")
    print(f"  活动能量: {len(activity_samples)} 条")
    print(f"  锻炼: {len(workouts)} 条")
    print(f"  体重: {len(weight_samples)} 条" + (f" (字段 '{w_key}')" if w_key else ""))
    print(f"  体脂率: {len(bodyfat_samples)} 条" + (f" (字段 '{bf_key}')" if bf_key else ""))

    # 结构自检
    if sleep_samples and not parse_date(sleep_samples[0].get("start", "")):
        print("\n  [警告] 睡眠样本 start 无法解析为日期, 请检查快捷指令属性选择")
    if activity_samples and not parse_date(activity_samples[0].get("start", "")):
        print("  [警告] 活动样本 start 无法解析为日期 (可能选成了来源名称)")

    sleep, sk1 = parse_sleep(sleep_samples)
    activity, sk2 = parse_activity(activity_samples)
    workout_data, sk3 = parse_workouts(workouts)
    weight, sk4 = parse_measure(weight_samples, WEIGHT_RANGE)
    bodyfat, sk5 = parse_measure(bodyfat_samples, BODYFAT_RANGE)

    # 剔除导出当天: 当天未过完, 数据不完整, 写入会误导
    # (第二天再导出时该日期会以完整值写入)
    if not args.include_today:
        today = bj_date_str(parse_date(data.get("exportDate", "")))
        if today:
            removed = []
            if today in sleep:
                sleep.pop(today)
                removed.append("睡眠")
            if today in activity:
                activity.pop(today)
                removed.append("活动能量")
            if today in weight:
                weight.pop(today)
                removed.append("体重")
            if today in bodyfat:
                bodyfat.pop(today)
                removed.append("体脂率")
            n0 = len(workout_data["strength"]) + len(workout_data["cardio"])
            workout_data["strength"] = [
                r for r in workout_data["strength"] if r["date"] != today
            ]
            workout_data["cardio"] = [
                r for r in workout_data["cardio"] if r["date"] != today
            ]
            if len(workout_data["strength"]) + len(workout_data["cardio"]) != n0:
                removed.append("锻炼")
            if removed:
                print(f"  [提示] 已剔除导出当天 {today} 的不完整数据"
                      f" ({'/'.join(removed)}), 次日会自动补全")

    print(f"\n解析结果:")
    print(f"  睡眠: {len(sleep)} 天 (跳过 {sk1} 条)")
    print(f"  活动能量: {len(activity)} 天 (跳过 {sk2} 条)")
    print(f"  体重: {len(weight)} 天 (跳过 {sk4} 条)")
    print(f"  体脂率: {len(bodyfat)} 天 (跳过 {sk5} 条)")
    print(f"  力量训练: {len(workout_data['strength'])} 条")
    print(f"  有氧训练: {len(workout_data['cardio'])} 条 (跳过 {sk3} 条)")

    # 每日汇总
    print(f"\n{'日期':<12} {'睡眠(h)':<8} {'活动(kcal)':<11} {'体重(kg)':<9} {'体脂(%)':<8} {'锻炼':<18}")
    print(f"{'-'*12} {'-'*8} {'-'*11} {'-'*9} {'-'*8} {'-'*18}")
    all_dates = sorted(
        set(sleep.keys()) | set(activity.keys())
        | set(weight.keys()) | set(bodyfat.keys())
    )
    for date in all_dates:
        sh = sleep.get(date, {}).get("sleep_hours", "-")
        wt = weight.get(date, {}).get("value", "-")
        bf = bodyfat.get(date, {}).get("value", "-")
        cal = activity.get(date, {}).get("calories", "-")
        w_str = ""
        for w in workout_data["strength"]:
            if w["date"] == date:
                w_str += f"力量({w['calories']}kcal) "
        for w in workout_data["cardio"]:
            if w["date"] == date:
                w_str += f"有氧({w['calories']}kcal) "
        if not w_str:
            w_str = "-"
        print(f"{date:<12} {sh:<8} {cal:<11} {wt:<9} {bf:<8} {w_str:<18}")

    # 验证
    print(f"\n数据验证:")
    warnings = validate(sleep, activity, workout_data, weight, bodyfat)
    if warnings:
        print(f"  {len(warnings)} 个警告:")
        for w in warnings:
            print(f"  ! {w}")
    else:
        print(f"  通过")

    # 写入
    if args.dry_run:
        print(f"\n--dry-run: 不写入 Excel")
    else:
        excel_path = find_excel(args.excel)
        merge_to_excel(excel_path, sleep, activity, workout_data,
                       weight, bodyfat)

    print(f"\n完成!")


if __name__ == "__main__":
    main()
